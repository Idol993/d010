import datetime
import json
import os

try:
    import openpyxl
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl is not installed. Excel export will be skipped.")

from .models import ReleaseRecord, ReleaseStatus, RollbackRecord, MonitoringMetrics
from .config import DATA_DIR


SUCCESS_STATUSES = {
    ReleaseStatus.FULLY_RELEASED,
    ReleaseStatus.STABLE_RESTORED,
    ReleaseStatus.MONITORING,
}


class RiskDashboard:

    def __init__(self):
        self.releases = []
        self.rollbacks = []
        self.monitoring_data = []

    def set_data(self, releases: list, rollbacks: list, monitoring_data: list):
        self.releases = releases
        self.rollbacks = rollbacks
        self.monitoring_data = monitoring_data

    def _calc_metrics(self, releases_in_group, rollbacks_in_group, monitoring_in_group):
        total_releases = len(releases_in_group)
        successful_releases = sum(
            1 for r in releases_in_group if r.status in SUCCESS_STATUSES
        )
        rollback_count = len(rollbacks_in_group)

        if monitoring_in_group:
            avg_loan_success_rate = sum(
                m.loan_success_rate for m in monitoring_in_group
            ) / len(monitoring_in_group)
        else:
            avg_loan_success_rate = 0.0

        alert_timestamps = [
            m.timestamp for m in monitoring_in_group
            if m.alert_triggered and m.timestamp is not None
        ]
        last_alert_time = max(alert_timestamps) if alert_timestamps else None

        return {
            "total_releases": total_releases,
            "successful_releases": successful_releases,
            "rollback_count": rollback_count,
            "avg_loan_success_rate": avg_loan_success_rate,
            "last_alert_time": last_alert_time,
        }

    def aggregate_by_enterprise(self) -> list:
        groups = {}
        for r in self.releases:
            key = r.enterprise_name or "未知企业"
            if key not in groups:
                groups[key] = {"releases": [], "release_ids": set()}
            groups[key]["releases"].append(r)
            groups[key]["release_ids"].add(r.id)

        result = []
        for name, group in groups.items():
            release_ids = group["release_ids"]
            group_rollbacks = [
                rb for rb in self.rollbacks if rb.release_id in release_ids
            ]
            group_monitoring = [
                m for m in self.monitoring_data if m.release_id in release_ids
            ]
            metrics = self._calc_metrics(
                group["releases"], group_rollbacks, group_monitoring
            )
            metrics["enterprise_name"] = name
            result.append(metrics)

        result.sort(key=lambda x: x["total_releases"], reverse=True)
        return result

    def aggregate_by_module(self) -> list:
        groups = {}
        for r in self.releases:
            key = r.industry_chain_module or "未知模块"
            if key not in groups:
                groups[key] = {"releases": [], "release_ids": set()}
            groups[key]["releases"].append(r)
            groups[key]["release_ids"].add(r.id)

        result = []
        for name, group in groups.items():
            release_ids = group["release_ids"]
            group_rollbacks = [
                rb for rb in self.rollbacks if rb.release_id in release_ids
            ]
            group_monitoring = [
                m for m in self.monitoring_data if m.release_id in release_ids
            ]
            metrics = self._calc_metrics(
                group["releases"], group_rollbacks, group_monitoring
            )
            metrics["industry_chain_module"] = name
            result.append(metrics)

        result.sort(key=lambda x: x["total_releases"], reverse=True)
        return result

    def aggregate_by_risk_level(self) -> list:
        groups = {}
        for r in self.releases:
            key = r.risk_level.value if r.risk_level else "未知风险"
            if key not in groups:
                groups[key] = {"releases": [], "release_ids": set()}
            groups[key]["releases"].append(r)
            groups[key]["release_ids"].add(r.id)

        result = []
        for name, group in groups.items():
            release_ids = group["release_ids"]
            group_rollbacks = [
                rb for rb in self.rollbacks if rb.release_id in release_ids
            ]
            group_monitoring = [
                m for m in self.monitoring_data if m.release_id in release_ids
            ]
            metrics = self._calc_metrics(
                group["releases"], group_rollbacks, group_monitoring
            )
            metrics["risk_level"] = name
            result.append(metrics)

        result.sort(key=lambda x: x["total_releases"], reverse=True)
        return result

    def print_dashboard(self):
        by_enterprise = self.aggregate_by_enterprise()
        by_module = self.aggregate_by_module()
        by_risk = self.aggregate_by_risk_level()

        self._print_table(
            "企业维度风险看板",
            ["企业名称", "发布总数", "成功发布", "回滚次数", "平均放款成功率(%)", "最近告警时间"],
            by_enterprise,
            [
                "enterprise_name",
                "total_releases",
                "successful_releases",
                "rollback_count",
                "avg_loan_success_rate",
                "last_alert_time",
            ],
        )
        print()

        self._print_table(
            "产业链模块维度风险看板",
            ["产业链模块", "发布总数", "成功发布", "回滚次数", "平均放款成功率(%)", "最近告警时间"],
            by_module,
            [
                "industry_chain_module",
                "total_releases",
                "successful_releases",
                "rollback_count",
                "avg_loan_success_rate",
                "last_alert_time",
            ],
        )
        print()

        self._print_table(
            "风险级别维度风险看板",
            ["风险级别", "发布总数", "成功发布", "回滚次数", "平均放款成功率(%)", "最近告警时间"],
            by_risk,
            [
                "risk_level",
                "total_releases",
                "successful_releases",
                "rollback_count",
                "avg_loan_success_rate",
                "last_alert_time",
            ],
        )

    def _print_table(self, title, headers, rows, keys):
        print("=" * 80)
        print(title)
        print("=" * 80)

        if not rows:
            print("(暂无数据)")
            return

        str_rows = []
        for row in rows:
            str_row = []
            for key in keys:
                val = row.get(key)
                if val is None:
                    str_row.append("-")
                elif isinstance(val, float):
                    str_row.append(f"{val:.2f}")
                elif isinstance(val, datetime.datetime):
                    str_row.append(val.strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    str_row.append(str(val))
            str_rows.append(str_row)

        col_widths = [len(h) for h in headers]
        for row in str_rows:
            for i, cell in enumerate(row):
                if len(cell) > col_widths[i]:
                    col_widths[i] = len(cell)

        def _pad(text, width):
            pad = width - len(text)
            return " " + text + " " * pad + " "

        sep = "+" + "+".join("-" * (w + 2) for w in col_widths) + "+"
        print(sep)
        header_line = "|" + "|".join(_pad(h, w) for h, w in zip(headers, col_widths)) + "|"
        print(header_line)
        print(sep)
        for row in str_rows:
            row_line = "|" + "|".join(_pad(c, w) for c, w in zip(row, col_widths)) + "|"
            print(row_line)
        print(sep)

    def export_excel(self, file_path: str = "") -> str:
        if not HAS_OPENPYXL:
            print("Warning: openpyxl is not available. Excel export skipped.")
            return ""

        if not file_path:
            now = datetime.datetime.now()
            filename = f"risk_dashboard_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(DATA_DIR, filename)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        wb = openpyxl.Workbook()

        by_enterprise = self.aggregate_by_enterprise()
        by_module = self.aggregate_by_module()
        by_risk = self.aggregate_by_risk_level()

        ws1 = wb.active
        ws1.title = "企业维度"
        headers1 = [
            "企业名称", "发布总数", "成功发布", "回滚次数",
            "平均放款成功率(%)", "最近告警时间",
        ]
        keys1 = [
            "enterprise_name", "total_releases", "successful_releases",
            "rollback_count", "avg_loan_success_rate", "last_alert_time",
        ]
        widths1 = [20, 12, 12, 12, 18, 22]
        self._write_sheet(ws1, headers1, keys1, by_enterprise, widths1)

        ws2 = wb.create_sheet("产业链模块维度")
        headers2 = [
            "产业链模块", "发布总数", "成功发布", "回滚次数",
            "平均放款成功率(%)", "最近告警时间",
        ]
        keys2 = [
            "industry_chain_module", "total_releases", "successful_releases",
            "rollback_count", "avg_loan_success_rate", "last_alert_time",
        ]
        widths2 = [22, 12, 12, 12, 18, 22]
        self._write_sheet(ws2, headers2, keys2, by_module, widths2)

        ws3 = wb.create_sheet("风险级别维度")
        headers3 = [
            "风险级别", "发布总数", "成功发布", "回滚次数",
            "平均放款成功率(%)", "最近告警时间",
        ]
        keys3 = [
            "risk_level", "total_releases", "successful_releases",
            "rollback_count", "avg_loan_success_rate", "last_alert_time",
        ]
        widths3 = [18, 12, 12, 12, 18, 22]
        self._write_sheet(ws3, headers3, keys3, by_risk, widths3)

        wb.save(file_path)
        return file_path

    def _write_sheet(self, ws, headers, keys, rows, widths):
        bold_font = Font(bold=True)
        for col, (header, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold_font
            ws.column_dimensions[chr(64 + col)].width = width

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, key in enumerate(keys, 1):
                val = row_data.get(key)
                if val is None:
                    cell_val = ""
                elif isinstance(val, float):
                    cell_val = round(val, 2)
                elif isinstance(val, datetime.datetime):
                    cell_val = val.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    cell_val = val
                ws.cell(row=row_idx, column=col_idx, value=cell_val)
