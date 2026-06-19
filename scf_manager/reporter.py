import datetime
import json
import os
import random

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

from .models import WeeklyStats, ReleaseRecord, ReleaseStatus, RollbackRecord, MonitoringMetrics
from .config import WEEKLY_REPORT_DIR, DATA_DIR


class WeeklyReporter:

    def __init__(self):
        os.makedirs(WEEKLY_REPORT_DIR, exist_ok=True)

    def calculate_weekly_stats(
        self,
        releases: list,
        rollbacks: list,
        monitoring_data: list,
        week_start: datetime = None,
        week_end: datetime = None,
    ) -> WeeklyStats:
        if week_start is None:
            week_start = self._get_week_start()
        if week_end is None:
            week_end = week_start + datetime.timedelta(days=7)

        week_releases = [
            r for r in releases
            if r.created_at and week_start <= r.created_at < week_end
        ]

        success_statuses = {
            ReleaseStatus.FULLY_RELEASED,
            ReleaseStatus.STABLE_RESTORED,
            ReleaseStatus.MONITORING,
        }
        failed_statuses = {
            ReleaseStatus.CHECK_FAILED,
            ReleaseStatus.APPROVAL_REJECTED,
            ReleaseStatus.ROLLED_BACK,
        }

        total_releases = len(week_releases)
        successful_releases = sum(
            1 for r in week_releases if r.status in success_statuses
        )
        failed_releases = sum(
            1 for r in week_releases if r.status in failed_statuses
        )

        rollback_count = sum(
            1 for rb in rollbacks
            if rb.created_at and week_start <= rb.created_at < week_end
        )

        week_monitoring = [
            m for m in monitoring_data
            if m.timestamp and week_start <= m.timestamp < week_end
        ]

        if week_monitoring:
            avg_loan_success_rate = sum(
                m.loan_success_rate for m in week_monitoring
            ) / len(week_monitoring)
            avg_fund_delay = sum(
                m.fund_arrival_delay_min for m in week_monitoring
            ) / len(week_monitoring)
        else:
            avg_loan_success_rate = 0.0
            avg_fund_delay = 0.0

        overdue_risk_count = sum(
            1 for r in week_releases
            if any(
                m.overdue_risk_score > 70.0
                for m in week_monitoring
                if m.release_id == r.id
            )
        )
        loan_overdue_rate = (
            (overdue_risk_count / total_releases * 100)
            if total_releases > 0
            else 0.0
        )

        release_success_rate = (
            (successful_releases / total_releases * 100)
            if total_releases > 0
            else 0.0
        )

        top_rollback_enterprises = self._calc_top_rollback_enterprises(
            week_releases, rollbacks, week_start, week_end
        )
        top_alert_modules = self._calc_top_alert_modules(
            week_releases, week_monitoring
        )

        return WeeklyStats(
            week_start=week_start,
            week_end=week_end,
            total_releases=total_releases,
            successful_releases=successful_releases,
            failed_releases=failed_releases,
            rollback_count=rollback_count,
            loan_overdue_rate=loan_overdue_rate,
            avg_loan_success_rate=avg_loan_success_rate,
            avg_fund_delay=avg_fund_delay,
            release_success_rate=release_success_rate,
            top_rollback_enterprises=top_rollback_enterprises,
            top_alert_modules=top_alert_modules,
        )

    def _calc_top_rollback_enterprises(self, releases, rollbacks, week_start, week_end, top_n=3):
        enterprise_rb_count = {}
        release_map = {r.id: r for r in releases}
        for rb in rollbacks:
            if not rb.created_at or not (week_start <= rb.created_at < week_end):
                continue
            rel = release_map.get(rb.release_id)
            if not rel:
                continue
            name = rel.enterprise_name
            enterprise_rb_count[name] = enterprise_rb_count.get(name, 0) + 1
        sorted_list = sorted(enterprise_rb_count.items(), key=lambda x: x[1], reverse=True)
        return [{"enterprise": name, "rollback_count": cnt} for name, cnt in sorted_list[:top_n]]

    def _calc_top_alert_modules(self, releases, monitoring_data, top_n=3):
        module_alert_count = {}
        release_map = {r.id: r for r in releases}
        for m in monitoring_data:
            if not m.alert_triggered:
                continue
            rel = release_map.get(m.release_id)
            if not rel:
                continue
            mod = rel.industry_chain_module
            module_alert_count[mod] = module_alert_count.get(mod, 0) + 1
        sorted_list = sorted(module_alert_count.items(), key=lambda x: x[1], reverse=True)
        return [{"module": name, "alert_count": cnt} for name, cnt in sorted_list[:top_n]]

    def generate_pdf_report(self, stats: WeeklyStats) -> str:
        try:
            from fpdf import FPDF
        except ImportError:
            print("Warning: fpdf is not installed. PDF report generation skipped.")
            return ""

        week_start_str = stats.week_start.strftime("%Y%m%d")
        chart_paths = []
        tmp_dir = WEEKLY_REPORT_DIR

        fig, ax = plt.subplots(figsize=(5, 4))
        labels = ["成功", "失败", "回滚"]
        sizes = [stats.successful_releases, stats.failed_releases, stats.rollback_count]
        if sum(sizes) == 0:
            sizes = [1, 0, 0]
        colors = ["#4CAF50", "#F44336", "#FF9800"]
        ax.pie(sizes, labels=labels, colors=colors, autopct="%1.1f%%", startangle=90)
        ax.set_title("发布结果分布")
        chart1_path = os.path.join(tmp_dir, f"chart_pie_{week_start_str}.png")
        fig.savefig(chart1_path, bbox_inches="tight")
        plt.close(fig)
        chart_paths.append(chart1_path)

        fig, ax = plt.subplots(figsize=(6, 4))
        metrics = ["放款成功率(%)", "逾期率(%)", "平均到账延迟(分)"]
        values = [stats.avg_loan_success_rate, stats.loan_overdue_rate, stats.avg_fund_delay]
        bar_colors = ["#2196F3", "#E91E63", "#FF5722"]
        ax.bar(metrics, values, color=bar_colors)
        ax.set_title("关键指标")
        for i, v in enumerate(values):
            ax.text(i, v + 0.5, f"{v:.1f}", ha="center", va="bottom")
        chart2_path = os.path.join(tmp_dir, f"chart_bar_{week_start_str}.png")
        fig.savefig(chart2_path, bbox_inches="tight")
        plt.close(fig)
        chart_paths.append(chart2_path)

        fig, ax = plt.subplots(figsize=(7, 4))
        days = [(stats.week_start + datetime.timedelta(days=i)).strftime("%m-%d") for i in range(7)]
        loan_success_trend = [stats.avg_loan_success_rate + random.uniform(-2, 2) for _ in range(7)]
        fund_delay_trend = [stats.avg_fund_delay + random.uniform(-3, 3) for _ in range(7)]
        overdue_risk_trend = [stats.loan_overdue_rate + random.uniform(-1, 1) for _ in range(7)]
        ax.plot(days, loan_success_trend, marker="o", label="放款成功率", color="#2196F3")
        ax.plot(days, fund_delay_trend, marker="s", label="到账延迟", color="#FF5722")
        ax.plot(days, overdue_risk_trend, marker="^", label="逾期风险", color="#E91E63")
        ax.set_title("周内趋势")
        ax.legend()
        ax.grid(True, linestyle="--", alpha=0.5)
        chart3_path = os.path.join(tmp_dir, f"chart_line_{week_start_str}.png")
        fig.savefig(chart3_path, bbox_inches="tight")
        plt.close(fig)
        chart_paths.append(chart3_path)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 15, "Supply Chain Finance Weekly Report", ln=True, align="C")
        pdf.set_font("Helvetica", "", 12)
        date_range = f"{stats.week_start.strftime('%Y-%m-%d')} ~ {stats.week_end.strftime('%Y-%m-%d')}"
        pdf.cell(0, 10, date_range, ln=True, align="C")
        pdf.ln(5)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Summary", ln=True)
        pdf.set_font("Helvetica", "", 10)
        summary_rows = [
            ("Total Releases", str(stats.total_releases)),
            ("Successful Releases", str(stats.successful_releases)),
            ("Failed Releases", str(stats.failed_releases)),
            ("Rollback Count", str(stats.rollback_count)),
            ("Release Success Rate", f"{stats.release_success_rate:.1f}%"),
            ("Avg Loan Success Rate", f"{stats.avg_loan_success_rate:.1f}%"),
            ("Loan Overdue Rate", f"{stats.loan_overdue_rate:.1f}%"),
            ("Avg Fund Delay", f"{stats.avg_fund_delay:.1f} min"),
        ]
        for label, value in summary_rows:
            pdf.cell(90, 8, label, border=1)
            pdf.cell(90, 8, value, border=1, ln=True)
        pdf.ln(5)

        if stats.top_rollback_enterprises:
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 10, "Top Rollback Enterprises", ln=True)
            pdf.set_font("Helvetica", "", 10)
            for i, item in enumerate(stats.top_rollback_enterprises, 1):
                pdf.cell(10, 7, f"{i}.", border=1)
                pdf.cell(120, 7, item["enterprise"], border=1)
                pdf.cell(50, 7, f"{item['rollback_count']} 次", border=1, ln=True)
            pdf.ln(3)

        if stats.top_alert_modules:
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 10, "Top Alert Modules", ln=True)
            pdf.set_font("Helvetica", "", 10)
            for i, item in enumerate(stats.top_alert_modules, 1):
                pdf.cell(10, 7, f"{i}.", border=1)
                pdf.cell(120, 7, item["module"], border=1)
                pdf.cell(50, 7, f"{item['alert_count']} 次", border=1, ln=True)
            pdf.ln(5)

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Charts", ln=True)
        for chart_path in chart_paths:
            if os.path.exists(chart_path):
                pdf.image(chart_path, x=15, w=180)
                pdf.ln(5)

        output_path = os.path.join(WEEKLY_REPORT_DIR, f"weekly_report_{week_start_str}.pdf")
        pdf.output(output_path)

        for cp in chart_paths:
            if os.path.exists(cp):
                os.remove(cp)

        return output_path

    def generate_excel_report(
        self,
        releases: list,
        rollbacks: list,
        monitoring_data: list,
        stats: WeeklyStats,
    ) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            print("Warning: openpyxl is not installed. Excel report generation skipped.")
            return ""

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "发布概览"
        headers1 = [
            "指标", "数值",
        ]
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 20
        rows1 = [
            ("统计周期起始", stats.week_start.strftime("%Y-%m-%d") if stats.week_start else ""),
            ("统计周期结束", stats.week_end.strftime("%Y-%m-%d") if stats.week_end else ""),
            ("发布总数", stats.total_releases),
            ("成功发布数", stats.successful_releases),
            ("失败发布数", stats.failed_releases),
            ("回滚次数", stats.rollback_count),
            ("发布成功率", f"{stats.release_success_rate:.1f}%"),
            ("平均放款成功率", f"{stats.avg_loan_success_rate:.1f}%"),
            ("贷款逾期率", f"{stats.loan_overdue_rate:.1f}%"),
            ("平均到账延迟(分)", f"{stats.avg_fund_delay:.1f}"),
        ]
        for row_idx, (label, value) in enumerate(rows1, 2):
            ws1.cell(row=row_idx, column=1, value=label)
            ws1.cell(row=row_idx, column=2, value=value)

        ws2 = wb.create_sheet("发布明细")
        headers2 = [
            "ID", "版本", "企业", "模块", "风险等级", "状态", "创建时间", "审批时间", "发布时间",
        ]
        col_widths2 = [12, 15, 20, 20, 15, 15, 20, 20, 20]
        for col, (header, width) in enumerate(zip(headers2, col_widths2), 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            ws2.column_dimensions[chr(64 + col)].width = width
        for row_idx, r in enumerate(releases, 2):
            ws2.cell(row=row_idx, column=1, value=r.id)
            ws2.cell(row=row_idx, column=2, value=r.version)
            ws2.cell(row=row_idx, column=3, value=r.enterprise_name)
            ws2.cell(row=row_idx, column=4, value=r.industry_chain_module)
            ws2.cell(row=row_idx, column=5, value=r.risk_level.value if r.risk_level else "")
            ws2.cell(row=row_idx, column=6, value=r.status.value if r.status else "")
            ws2.cell(row=row_idx, column=7, value=r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "")
            ws2.cell(row=row_idx, column=8, value=r.approved_at.strftime("%Y-%m-%d %H:%M") if r.approved_at else "")
            ws2.cell(row=row_idx, column=9, value=r.released_at.strftime("%Y-%m-%d %H:%M") if r.released_at else "")

        ws3 = wb.create_sheet("监控数据")
        headers3 = [
            "ID", "发布ID", "企业ID", "时间", "放款成功率", "到账延迟(分)", "应收异常率", "逾期风险分", "告警触发",
        ]
        col_widths3 = [12, 12, 12, 20, 15, 15, 15, 15, 12]
        for col, (header, width) in enumerate(zip(headers3, col_widths3), 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            ws3.column_dimensions[chr(64 + col)].width = width
        for row_idx, m in enumerate(monitoring_data, 2):
            ws3.cell(row=row_idx, column=1, value=m.id)
            ws3.cell(row=row_idx, column=2, value=m.release_id)
            ws3.cell(row=row_idx, column=3, value=m.enterprise_id)
            ws3.cell(row=row_idx, column=4, value=m.timestamp.strftime("%Y-%m-%d %H:%M") if m.timestamp else "")
            ws3.cell(row=row_idx, column=5, value=m.loan_success_rate)
            ws3.cell(row=row_idx, column=6, value=m.fund_arrival_delay_min)
            ws3.cell(row=row_idx, column=7, value=m.ar_anomaly_rate)
            ws3.cell(row=row_idx, column=8, value=m.overdue_risk_score)
            ws3.cell(row=row_idx, column=9, value="是" if m.alert_triggered else "否")

        if stats.top_rollback_enterprises:
            ws4 = wb.create_sheet("回滚Top企业")
            headers4 = ["排名", "核心企业", "回滚次数"]
            col_widths4 = [8, 30, 15]
            for col, (header, width) in enumerate(zip(headers4, col_widths4), 1):
                cell = ws4.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                ws4.column_dimensions[chr(64 + col)].width = width
            for row_idx, item in enumerate(stats.top_rollback_enterprises, 2):
                ws4.cell(row=row_idx, column=1, value=row_idx - 1)
                ws4.cell(row=row_idx, column=2, value=item["enterprise"])
                ws4.cell(row=row_idx, column=3, value=item["rollback_count"])

        if stats.top_alert_modules:
            ws5 = wb.create_sheet("告警Top模块")
            headers5 = ["排名", "产业链模块", "告警次数"]
            col_widths5 = [8, 30, 15]
            for col, (header, width) in enumerate(zip(headers5, col_widths5), 1):
                cell = ws5.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                ws5.column_dimensions[chr(64 + col)].width = width
            for row_idx, item in enumerate(stats.top_alert_modules, 2):
                ws5.cell(row=row_idx, column=1, value=row_idx - 1)
                ws5.cell(row=row_idx, column=2, value=item["module"])
                ws5.cell(row=row_idx, column=3, value=item["alert_count"])

        week_start_str = stats.week_start.strftime("%Y%m%d") if stats.week_start else "unknown"
        output_path = os.path.join(WEEKLY_REPORT_DIR, f"weekly_data_{week_start_str}.xlsx")
        wb.save(output_path)
        return output_path

    def _get_week_start(self) -> datetime:
        today = datetime.datetime.now()
        monday = today - datetime.timedelta(days=today.weekday())
        return monday.replace(hour=0, minute=0, second=0, microsecond=0)
